"""
从原始 Excel 设备物料清单补充本地数据库 supplier 的产品数据

执行:
  cd backend
  python enrich_suppliers_from_excel.py

作用:
  - 读取 清单-设备科常用无编码的物料.xlsx
  - 按供应商名映射到 DB 的 equipment 品类供应商
  - 补充 products / description / capabilities 到 attributes JSONB
"""

import json, os, re
from collections import defaultdict
from typing import Optional

import psycopg2
from psycopg2.extras import RealDictCursor

DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL is required. Set it in your environment or backend/.env; never hard-code database credentials.")

EXCEL_DIR = os.path.join(os.path.dirname(__file__),
    "..", "AW_ [EXT] Request for Project Background – AI Procurement Agent for Sourcing & Cost Comparison"
)
EXCEL_PATH = os.path.join(EXCEL_DIR, "清单-设备科常用无编码的物料.xlsx")


def s(v):
    if v is None: return ""
    return str(v).strip()


def load_excel_products():
    """从 Excel 读取所有供应商的产品清单"""
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH)

    supplier_products = defaultdict(list)

    # ---- Sheet 1: 楼下物料 (MRO/设备备件) ----
    ws = wb["楼下物料"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        part_name = s(row[0])
        desc = s(row[1])
        model = s(row[2])
        artno = s(row[3])
        manufacturer = s(row[4])
        supplier = s(row[5])

        if not supplier or not part_name:
            continue

        # 构建简洁的产品名（供词法匹配用）
        brief = part_name
        if model:
            brief += f" {model}"
        supplier_products[supplier].append({
            "name": part_name,
            "model": model,
            "description": desc,
            "articleNo": artno,
            "manufacturer": manufacturer,
            "brief": brief,
        })

    # ---- Sheet 2: FM system ----
    ws2 = wb["FM system"]
    for row in ws2.iter_rows(min_row=2, values_only=True):
        part_name = s(row[0])
        desc = s(row[1])
        model = s(row[2])
        artno = s(row[3])
        manufacturer = s(row[4])
        supplier = s(row[5])

        if not supplier or not part_name:
            continue
        brief = part_name
        if model:
            brief += f" {model}"
        supplier_products[supplier].append({
            "name": part_name,
            "model": model,
            "description": desc,
            "articleNo": artno,
            "manufacturer": manufacturer,
            "brief": brief,
        })

    wb.close()
    return supplier_products


# ============================================================
# 供应商映射: Excel中的供应商名 → DB供应商id
# ============================================================
DB_SUPPLIERS = {}  # 将在运行时从 DB 加载

# 分组映射: 多个 Excel 供应商名 → 同一个 DB id
# 相同 DB id 的产品会被合并
SUPPLIER_MAP: dict[int, list[str]] = {
    124: ["Festo", "FESTO"],           # Festo AG & Co. KG
    125: ["Bieler"],                    # Bieler Industriebedarf
    93:  ["SMC"],
    89:  ["Siemens AG", "Siemen AG"],  # Siemens AG (含 typo)
    90:  ["RS-Components"],
    94:  ["Schmalz"],
    95:  ["SF-Filter"],
    91:  ["Weidmüller"],
    86:  ["Unielektro"],
    88:  ["Fischer Zander"],
    97:  ["FM Systeme GmbH", "FM-SYSTEM"],
    96:  ["FM-SYSTEM/Weinmann"],
    98:  ["Weinmann"],
    99:  ["Rala"],
    100: ["Lottor"],
    92:  ["ABL"],
}


def build_capabilities(supplier_name: str, products: list) -> list:
    """根据产品类型推断供应商能力"""
    # 按产品名称归类
    categories_seen = set()
    for p in products:
        name = p["name"].lower()
        if any(kw in name for kw in ["密封", "dicht", "teflon"]):
            categories_seen.add("dichtungstechnik")
        if any(kw in name for kw in ["接头", "steck", "schlauch", "pneumatik", "ventil", "zylinder", "druck"]):
            categories_seen.add("pneumatik")
        if any(kw in name for kw in ["schaltschrank", "schütz", "schalter", "sicherung", "relais", "trenner", "fi-"]):
            categories_seen.add("elektroinstallation")
        if any(kw in name for kw in ["kabel", "leitung", "stecker", "dose", "verbinder"]):
            categories_seen.add("kabeltechnik")
        if any(kw in name for kw in ["lager", "führung", "schiene", "lineartechnik", "profil"]):
            categories_seen.add("lineartechnik")
        if any(kw in name for kw in ["filter", "sieb"]):
            categories_seen.add("filtertechnik")
        if any(kw in name for kw in ["blech", "stahl", "aluminium", "metall", "platte"]):
            categories_seen.add("metallverarbeitung")
        if any(kw in name for kw in ["pumpe", "motor", "getriebe"]):
            categories_seen.add("antriebstechnik")
        if any(kw in name for kw in ["schraub", "mutter", "scheibe", "dübel", "befestigung"]):
            categories_seen.add("befestigungstechnik")
        if any(kw in name for kw in ["rohr", "fitting", "kupplung"]):
            categories_seen.add("rohrleitungstechnik")

    caps = [
        f"Lieferant für {', '.join(sorted(categories_seen)[:3])}" if categories_seen else "Industriebedarf",
    ]
    unique_manufacturers = set(p.get("manufacturer", "") for p in products if p.get("manufacturer"))
    if unique_manufacturers:
        brands = [m for m in unique_manufacturers if len(m) > 2][:3]
        if brands:
            caps.append(f"Vertriebspartner von {' / '.join(brands)}")
    caps.append(f"{len(products)}+ Produkte auf Lager")

    # Generic equipment capabilities
    if any(kw in supplier_name.lower() for kw in ["elektro", "siemens", "rs-"]):
        caps.append("Elektrotechnik & Automation")
    return caps


def build_description(supplier_name: str, products: list, caps: list) -> str:
    """生成德文描述"""
    product_summary = defaultdict(int)
    for p in products[:20]:
        product_summary[p["name"]] += 1

    top_categories = sorted(set(p["name"] for p in products))[:8]
    top_manufacturers = sorted(set(
        p.get("manufacturer", "") for p in products if p.get("manufacturer") and len(p.get("manufacturer", "")) > 2
    ))[:3]

    base = f"{supplier_name} ist ein zuverlässiger Lieferant für Industriebedarf, Maschinenkomponenten und Betriebsmittel."

    if top_manufacturers:
        base += f" Als Vertriebspartner von {' / '.join(top_manufacturers)} bietet das Unternehmen eine breite Palette an technischen Produkten."

    base += f" Der Katalog umfasst {len(products)} Produkte aus den Bereichen {', '.join(caps[:3])}."

    return base


def main():
    # 连接 DB
    print("📡 连接 Supabase...")
    conn = psycopg2.connect(DATABASE_URL)
    cur = conn.cursor(cursor_factory=RealDictCursor)

    # 加载 Excel 数据
    print("📂 读取 Excel...")
    excel_products = load_excel_products()
    print(f"  Excel 中共 {sum(len(v) for v in excel_products.values())} 个产品记录")
    print(f"  涉及 {len(excel_products)} 家供应商")

    # 逐个供应商更新
    updated = 0
    skipped = []
    total_products_added = 0

    for db_id, excel_names in SUPPLIER_MAP.items():
        # 合并同名供应商下多个 Excel 来源的产品
        all_raw = []
        for en in excel_names:
            all_raw.extend(excel_products.get(en, []))

        if not all_raw:
            src_label = " / ".join(excel_names)
            skipped.append((src_label, db_id, "Excel 中无产品数据"))
            continue

        # 去重: 用 产品名+型号+货号 做 dedup key（保留描述差异以丰富词法匹配）
        seen = set()
        unique_products = []
        for p in all_raw:
            # key 包含 articleNo 以区分同产品名不同规格
            key = p["brief"] + "|" + p["articleNo"]
            if key not in seen:
                seen.add(key)
                unique_products.append(p)

        # 构建 attributes 用的产品列表: 包含名称+描述摘要, 供词法匹配
        unique_briefs = []
        for p in unique_products:
            entry = p["brief"]
            if p.get("description"):
                desc_short = p["description"][:60]
                entry += f" - {desc_short}"
            if p.get("articleNo"):
                entry += f" [Art. {p['articleNo']}]"
            unique_briefs.append(entry)
        src_label = " / ".join(excel_names)
        caps = build_capabilities(src_label, unique_products)
        desc = build_description(src_label, unique_products, caps)

        new_attrs = {
            "category": "equipment",
            "products": unique_briefs,
            "description": desc,
            "capabilities": caps,
            "certifications": [],
            "city": "",
        }

        # 获取当前 attributes，保留已有数据
        cur.execute("SELECT attributes FROM supplier WHERE id = %s", (db_id,))
        row = cur.fetchone()
        if not row:
            skipped.append((src_label, db_id, "DB 中不存在"))
            continue

        current_attrs = row["attributes"] or {}

        # 保留 id 等字段，覆盖产品数据
        current_attrs["category"] = "equipment"
        current_attrs["products"] = unique_briefs
        current_attrs["description"] = desc
        current_attrs["capabilities"] = caps

        # 如果之前没有 certifications，保留空列表
        if "certifications" not in current_attrs:
            current_attrs["certifications"] = []

        # 更新 DB
        cur.execute(
            "UPDATE supplier SET attributes = %s WHERE id = %s",
            (json.dumps(current_attrs), db_id),
        )
        updated += 1
        total_products_added += len(unique_briefs)
        src_label = " / ".join(excel_names)
        print(f"  ✅ [id={db_id}] {src_label:<30} → {len(unique_briefs)} 个产品")

    # 对于没有 Excel 数据的基础供应商，补充简单描述
    basic_no_excel = [("Conrad Electronic", 81), ("S-POLYTEC GmbH/First Bond", 101)]
    for name, db_id in basic_no_excel:
        cur.execute("SELECT attributes FROM supplier WHERE id = %s", (db_id,))
        row = cur.fetchone()
        if not row:
            continue
        attrs = row["attributes"] or {}
        if attrs.get("products") and len(attrs.get("products", [])) > 0:
            continue  # 已有产品
        # 只补充描述
        if not attrs.get("description"):
            attrs["description"] = f"{name} ist ein etablierter Lieferant für Industriebedarf und technische Komponenten."
            attrs["capabilities"] = ["Industriebedarf"]
            cur.execute("UPDATE supplier SET attributes = %s WHERE id = %s", (json.dumps(attrs), db_id))
            print(f"  📝 [id={db_id}] {name:<30} → 补充了基础描述")

    conn.commit()

    print(f"\n{'='*60}")
    print(f"📊 汇总:")
    print(f"  更新: {updated} 家供应商, 共添加 {total_products_added} 个产品")
    print(f"  跳过: {len(skipped)} 家")
    for name, db_id, reason in skipped:
        print(f"    ⏭️  {name} (id={db_id}): {reason}")

    cur.close()
    conn.close()
    print("✅ 完成！")


if __name__ == "__main__":
    main()
